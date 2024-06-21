import pyautogui as pag
import time
import openpyxl
import pyperclip
import os
from PyQt5.QtWidgets import QMessageBox


class AutoWork:
    def __init__(self, filename):
        """打开filename文件，找到存储文件的活动页"""
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.sheet = self.wb.active
        # 检测是否合法
        self.data_check()

    def data_check(self):
        """检查数据是否合法"""
        check_result = True

        # 检查列数是否小于2，如果小于说明文件中没有写入内容，检查失败。
        if self.sheet.max_row < 2:
            QMessageBox.critical(None, '错误提示', '没有检测到数据')
            check_result = False
            return check_result

        i = 2
        while i <= self.sheet.max_row:
            # 检查指令类型是否为数字
            cmdvalue = self.sheet.cell(row=i, column=1).value
            if not isinstance(cmdvalue, int) or cmdvalue not in (1,2,3,4,5,6,7,8,9,10,11):
                QMessageBox.critical(None, '错误提示', '指令类型不是数字，或者输入的执行范围不在1-11')
                check_result = False
                return check_result
            # 移动到指定的坐标
            if cmdvalue == 1:
                data = self.sheet.cell(row=i, column=2).value
                if len(data) != 2:
                    QMessageBox.critical(None, '错误提示', f"第{i}行第2列数据有问题"+'请输入数字二元组')
                try:
                    x, y = data.split(',')
                    x = int(x)
                    y = int(y)
                except:
                    QMessageBox.critical(None, '错误提示', f"第{i}行第2列数据有问题"+'请输入数字二元组')
                    check_result = False
                    return check_result
            # 单击左键
            if cmdvalue == 2:
                pass
            # 单击右键
            if cmdvalue == 3:
                pass
            # 单击坐标
            if cmdvalue == 4:
                data = self.sheet.cell(row=i, column=2).value
                if len(data) != 2:
                    QMessageBox.critical(None, '错误提示', f"第{i}行第2列数据有问题"+'请输入数字二元组')
                try:
                    x, y = data.split(',')
                    x = int(x)
                    y = int(y)
                except:
                    QMessageBox.critical(None, '错误提示', f"第{i}行第2列数据有问题"+'请输入数字二元组')
                    check_result = False
                    return check_result
            # 单击图片
            if cmdvalue == 5:
                if not isinstance(self.sheet.cell(row=i, column=3).value, str):
                    QMessageBox.critical(None, '错误提示', f"第{i}行第3列数据有问题"+'请输入图片名称')
                    check_result = False
                    return check_result
            # 滑动滚轮
            if cmdvalue == 6:
                if not isinstance(self.sheet.cell(row=i, column=4).value, int):
                    QMessageBox.critical(None, '错误提示', f"第{i}行第4列数据有问题" + '请输入滚动距离')
                    check_result = False
                    return check_result
            # 等待时间
            if cmdvalue == 7:
                if not isinstance(self.sheet.cell(row=i, column=5).value, int):
                    QMessageBox.critical(None, '错误提示', f"第{i}行第5列数据有问题" + '请输入等待时间')
                    check_result = False
                    return check_result
            # 输入内容
            if cmdvalue == 8:
                if not isinstance(self.sheet.cell(row=i, column=6).value, str):
                    QMessageBox.critical(None, '错误提示', f"第{i}行第6列数据有问题" + '请输入内容')
                    check_result = False
                    return check_result
            # 指定循环操作
            if cmdvalue == 9:
                pass

            # 输入内容
            if cmdvalue == 4:
                if not isinstance(self.sheet.cell(row=i, column=2).value, str) or type(
                        self.sheet.cell(row=i, column=5).value) != type(None) or (
                        type(self.sheet.cell(row=i, column=5).value) != type(None) and (
                not isinstance(self.sheet.cell(row=i, column=5).value, int))):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False


            # 回车
            elif cmdvalue == 7:
                if type(self.sheet.cell(row=i, column=5).value) != type(None) and (
                not isinstance(self.sheet.cell(row=i, column=5).value, int)):
                    print("第%s行第2列数据有问题" % i)
                    check_result = False
            # 移动鼠标
            elif cmdvalue == 8:
                if not isinstance(self.sheet.cell(row=i, column=3).value, int) or not isinstance(
                        self.sheet.cell(row=i, column=4).value, int) or (
                        type(self.sheet.cell(row=i, column=5).value) != type(None) and (
                not isinstance(self.sheet.cell(row=i, column=5).value, int))):
                    print("第%s行第2列,3列，4列数据有问题" % i)

            i += 1
            return check_result